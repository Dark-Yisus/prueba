import os
import sys
import logging
import tempfile
import shutil
import time
import webbrowser
from datetime import datetime
from threading import Timer
from concurrent.futures import ThreadPoolExecutor, as_completed
import io

import flask
from flask import Flask, jsonify, request, render_template, send_file
from flask_cors import CORS
from werkzeug.middleware.proxy_fix import ProxyFix

import requests
import pandas as pd
import pymongo
from pymongo import MongoClient
from bs4 import BeautifulSoup
import re
import urllib3

# Disable SSL warnings
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# Determine if we're running in a bundle
if getattr(sys, 'frozen', False):
    template_folder = os.path.join(sys._MEIPASS, 'templates')
    app = Flask(__name__, template_folder=template_folder)
else:
    app = Flask(__name__)

CORS(app, resources={r"/api/*": {"origins": "*"}})
app.wsgi_app = ProxyFix(app.wsgi_app)

# Configure logging
log_file = os.path.join(os.path.expanduser("~"), 'mercadolibre_search.log')
logging.basicConfig(
    filename=log_file,
    level=logging.DEBUG,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# Add console handler for development
console_handler = logging.StreamHandler()
console_handler.setLevel(logging.DEBUG)
logger.addHandler(console_handler)

# MongoDB configuration
MONGO_URI = 'mongodb://localhost:27017/'
DB_NAME = 'multi_ecommerce_db'
COLLECTION_NAME = 'productos'

def open_browser():
    """Open the default web browser to the application URL"""
    webbrowser.open('http://127.0.0.1:5000/')

def get_mongodb_connection():
    try:
        client = MongoClient(MONGO_URI, serverSelectionTimeoutMS=5000)
        client.server_info()
        db = client[DB_NAME]
        return client, db
    except Exception as e:
        logger.error(f"Error connecting to MongoDB: {e}")
        return None, None

def extraer_cantidad_vendidos(url, headers):
    try:
        # Aumentar timeout y manejar certificados SSL
        response = requests.get(url, headers=headers, timeout=10, verify=False)
        
        if response.status_code != 200:
            logger.warning(f"No se pudo obtener la URL {url}. Código de estado: {response.status_code}")
            return 0
        
        soup = BeautifulSoup(response.text, 'html.parser')
        
        # Patrones para buscar cantidad de vendidos
        patrones = [
            # Patrón para español
            r'(\d+(?:\.\d+)?)\s*(?:vendidos?)',
            # Patrón para inglés
            r'(\d+(?:\.\d+)?)\s*(?:sold)',
            # Patrón más genérico
            r'(\d+(?:\.\d+)?)\s*(?:vendido|ventas)'
        ]
        
        # Buscar en todo el texto
        texto_completo = soup.get_text()
        
        for patron in patrones:
            match = re.search(patron, texto_completo, re.IGNORECASE)
            if match:
                try:
                    cantidad = int(float(match.group(1)))
                    logger.info(f"Productos vendidos encontrados: {cantidad}")
                    return cantidad
                except ValueError:
                    continue
        
        # Si no se encuentra nada, intentar extraer de elementos HTML específicos
        elementos_posibles = [
            soup.find('span', class_='ui-pdp-color--GREEN ui-pdp-family--REGULAR'),
            soup.find('div', class_='ui-pdp-color--GREEN ui-pdp-family--REGULAR')
        ]
        
        for elemento in elementos_posibles:
            if elemento and elemento.text:
                match = re.search(r'(\d+)', elemento.text)
                if match:
                    try:
                        cantidad = int(match.group(1))
                        logger.info(f"Productos vendidos extraídos de elemento HTML: {cantidad}")
                        return cantidad
                    except ValueError:
                        continue
        
        logger.warning(f"No se encontraron productos vendidos en {url}")
        return 0
    
    except Exception as e:
        logger.error(f"Error al extraer productos vendidos de {url}: {e}")
        return 0
    
def buscar_producto_api(producto, offset, batch_size):
    try:
        url = "https://api.mercadolibre.com/sites/MLM/search"
        params = {
            "q": producto,
            "offset": offset,
            "limit": batch_size
        }
        
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
        }
        
        response = requests.get(url, params=params, headers=headers, timeout=10)
        response.raise_for_status()
        
        resultado = response.json()
        
        if not resultado or 'results' not in resultado:
            logger.warning(f"No se encontraron resultados para {producto}")
            return None
        
        return resultado
    
    except requests.RequestException as e:
        logger.error(f"Error en la solicitud a la API de MercadoLibre: {e}")
        return None
    except Exception as e:
        logger.error(f"Error inesperado en buscar_producto_api: {e}")
        return None

def generate_excel(productos):
    """Generar un archivo Excel con los productos y sus imágenes individuales en formato PNG"""
    try:
        # Crear un DataFrame con las columnas seleccionadas
        df = pd.DataFrame(productos)
        
        # Mapa de columnas actualizado
        columns_map = {
            'producto': 'Producto',
            'precio_original': 'Precio Original',
            'precio_con_descuento': 'Precio con Descuento',
            'descuento': 'Descuento (%)',
            'vendedor': 'Vendedor',
            'estado_producto': 'Estado del Producto',
            'cantida_vendido': 'Cantidad Vendida',
            'cuotas': 'Cuotas Disponibles',
            'meses_intereses': 'Meses sin Intereses',
            'envio_gratis': 'Envío Gratis',
            'cantidad_disponible': 'Cantidad Disponible',
            'url_producto': 'URL del Producto',
            'categoria': 'Categoría'
        }
        
        # Crear directorio temporal para imágenes
        temp_image_dir = tempfile.mkdtemp()
        
        # Crear el archivo Excel en memoria
        output = io.BytesIO()
        
        # Usar openpyxl en lugar de xlsxwriter para mejor compatibilidad
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Preparar DataFrame
            df_export = df[list(columns_map.keys())].copy()
            df_export = df_export.rename(columns=columns_map)
            
            # Exportar a Excel
            df_export.to_excel(writer, sheet_name='Productos', index=False, startrow=1)
            
            workbook = writer.book
            worksheet = writer.sheets['Productos']
            
            # Configurar formato de encabezados (simplificado para openpyxl)
            from openpyxl.styles import Font, PatternFill, Alignment
            
            header_font = Font(bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='1E88E5', end_color='1E88E5', fill_type='solid')
            header_alignment = Alignment(horizontal='center', vertical='center')
            
            # Aplicar formato a los encabezados
            for col_num, value in enumerate(df_export.columns.values, 1):
                cell = worksheet.cell(row=1, column=col_num)
                cell.value = value
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                
                # Ajustar ancho de columna
                worksheet.column_dimensions[cell.column_letter].width = 20
            
            # Añadir columna de imágenes
            imagen_col = len(df_export.columns) + 1
            imagen_cell = worksheet.cell(row=1, column=imagen_col)
            imagen_cell.value = 'Imagen'
            imagen_cell.font = header_font
            imagen_cell.fill = header_fill
            imagen_cell.alignment = header_alignment
            worksheet.column_dimensions[imagen_cell.column_letter].width = 20
            
            # Procesar cada producto (sin imágenes para evitar problemas de compatibilidad)
            for idx, row in df.iterrows():
                try:
                    # Configurar altura de fila
                    worksheet.row_dimensions[idx + 2].height = 60
                    
                    # Solo agregar el URL de la imagen como texto
                    imagen_url = row.get('imagen_url', 'N/A')
                    imagen_cell = worksheet.cell(row=idx + 2, column=imagen_col)
                    imagen_cell.value = imagen_url if pd.notna(imagen_url) else 'Sin imagen'
                    
                except Exception as e:
                    logger.error(f"Error procesando fila {idx}: {e}")
                    continue
        
        # Limpiar directorio temporal
        shutil.rmtree(temp_image_dir, ignore_errors=True)
        
        output.seek(0)
        return output
    
    except Exception as e:
        logger.error(f"Error generando archivo Excel: {e}")
        return None

def guardar_productos_en_db(productos):
    try:
        # Intentar establecer conexión con MongoDB
        client, db = get_mongodb_connection()
        if not client or not db:
            logger.error("No se pudo establecer conexión con MongoDB")
            return False
        
        # Obtener la colección
        collection = db[COLLECTION_NAME]
        
        # Usar bulk write para mejor rendimiento
        bulk_operations = []
        for producto in productos:
            # Crear una operación de upsert basada en la URL del producto para evitar duplicados
            bulk_operations.append(
                pymongo.UpdateOne(
                    {"url_producto": producto.get("url_producto", "")},  # Identificador único
                    {"$set": producto},  # Actualizar o insertar todo el documento del producto
                    upsert=True  # Insertar si no existe
                )
            )
        
        # Ejecutar operaciones en bloque
        if bulk_operations:
            result = collection.bulk_write(bulk_operations)
            logger.info(f"Productos guardados en MongoDB: {result.upserted_count} insertados, {result.modified_count} actualizados")
            return True
        
        return False
    
    except Exception as e:
        logger.error(f"Error guardando productos en MongoDB: {e}")
        return False

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/api/search', methods=['POST'])
def mercadoLibre():
    try:
        logger.debug("Solicitud de búsqueda recibida")
        data = request.get_json()
        if not data or "producto" not in data:
            logger.warning("Datos de solicitud inválidos")
            return jsonify({"error": "Datos del producto no proporcionados"}), 400

        # Headers para simular navegador
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36',
            'Accept-Language': 'es-ES,es;q=0.9',
            'Accept-Encoding': 'gzip, deflate, br',
            'Connection': 'keep-alive'
        }

        # Get initial count of total results
        resultado_inicial = buscar_producto_api(data["producto"], 0, 1)
        if not resultado_inicial:
            logger.warning("No se encontraron resultados de la API")
            return jsonify({"error": "No se encontraron productos o ocurrió un error en la búsqueda"}), 404

        # Establecer límite a 2000 productos
        total_resultados = min(resultado_inicial['paging']['total'], 2000)
        batch_size = 50
        productos_procesados = []
        
        # Usar ThreadPoolExecutor para procesamiento en paralelo
        with ThreadPoolExecutor(max_workers=10) as executor:
            # Crear futures para todos los offsets
            future_to_offset = {
                executor.submit(buscar_producto_api, data["producto"], offset, batch_size): offset
                for offset in range(0, total_resultados, batch_size)
            }
            
            # Procesar lotes de resultados
            for future in as_completed(future_to_offset):
                offset = future_to_offset[future]
                try:
                    resultados_lote = future.result()
                    if resultados_lote and 'results' in resultados_lote:
                        # Preparar futures para vendidos
                        futures_vendidos = {}
                        
                        for item in resultados_lote['results']:
                            try:
                                producto = {
                                    "producto": item.get('title', 'N/A'),
                                    "plataforma": "mercadolibre",
                                    "precio_original": float(item.get('price', 0)),
                                    "precio_con_descuento": float(item.get('price', item.get('original_price', 0))),
                                    "descuento": float(item.get('discount_percentage', 0)),
                                    "vendedor": item.get('seller', {}).get('nickname', 'N/A'),
                                    "cuotas": int(item.get('installments', {}).get('quantity', 0)),
                                    "meses_intereses": int(item.get('installments', {}).get('months', 0)),
                                    "envio_gratis": 'Gratis' if item.get('shipping', {}).get('free_shipping', False) else 'N/A',
                                    "estado_producto": 'nuevo' if item.get('condition') == 'new' else 'usado' if item.get('condition') == 'used' else 'N/A',
                                    "cantidad_disponible": int(item.get('available_quantity', 0)),
                                    "url_producto": item.get('permalink', 'N/A'),
                                    "imagen_url": item.get('thumbnail', 'N/A'),
                                    "fecha_extraccion": datetime.now(),
                                    "categoria": item.get('category_id', 'N/A')
                                }
                                
                                # Agregar future para extraer vendidos
                                future_vendidos = executor.submit(
                                    extraer_cantidad_vendidos, 
                                    producto['url_producto'], 
                                    headers
                                )
                                futures_vendidos[future_vendidos] = producto
                            
                            except Exception as e:
                                logger.error(f"Error procesando producto individual: {e}")
                        
                        # Procesar resultados de vendidos
                        for future_venta in as_completed(futures_vendidos):
                            try:
                                vendidos = future_venta.result()
                                producto = futures_vendidos[future_venta]
                                producto['cantida_vendido'] = vendidos
                                productos_procesados.append(producto)
                            except Exception as e:
                                logger.error(f"Error asignando vendidos: {e}")
                        
                    # Agregar pequeño retardo entre lotes
                    time.sleep(0.2)
                
                except Exception as e:
                    logger.error(f"Error procesando lote en offset {offset}: {e}")
        
        if not productos_procesados:
            logger.warning("No se pudieron procesar productos")
            return jsonify({"error": "No se encontraron productos o ocurrió un error en la búsqueda"}), 404

        # Guardar en base de datos
        db_success = guardar_productos_en_db(productos_procesados)
            
        return jsonify({
            "success": True,
            "data": productos_procesados,
            "total": len(productos_procesados),
            "db_saved": db_success
        })
            
    except Exception as e:
        logger.error(f"Error inesperado en mercadoLibre: {str(e)}", exc_info=True)
        return jsonify({"error": "Error interno del servidor", "details": str(e)}), 500

@app.route('/api/download-excel', methods=['POST'])
def download_excel():
    try:
        data = request.get_json()
        if not data or "productos" not in data:
            return jsonify({"error": "No se proporcionaron datos de productos"}), 400
            
        excel_file = generate_excel(data["productos"])
        if not excel_file:
            return jsonify({"error": "Error generando archivo Excel"}), 500
            
        return send_file(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'productos_mercadolibre_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        )
        
    except Exception as e:
        logger.error(f"Error en download_excel: {e}")
        return jsonify({"error": "Error generando archivo Excel", "details": str(e)}), 500

if __name__ == "__main__":
    # Open browser after a short delay
    Timer(1.5, open_browser).start()
    
    # Get port that's not in use
    port = 5000
    while True:
        try:
            app.run(host="127.0.0.1", port=port, debug=False)
            break
        except OSError:
            port += 1